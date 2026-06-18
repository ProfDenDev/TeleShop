# app/imports/parsers/xlsx_parser.py
# ver 3.0
# ver 4.0

"""
Парсер XLSX для TELESHOP.

Поддерживаемый лист:

    Товары

Поддерживаемые колонки:

    sku
    title

    short_description
    full_description
    content_language

    category_slug
    subcategory_slug

    brand_name

    manufacturer_name
    manufacturer_sku

    barcode
    serial_number

    condition

    price_type
    currency
    price

    quantity

    weight_g
    dimensions

    location_area

    video_url

    attributes_json

    photo_files

Важно:

price может содержать:

    500

    500-2000

    10.5

    10,5

    0.55-0.83

Фото:

    IMG_001.jpg

или

    IMG_001.jpg|IMG_002.jpg

или

    IMG_001.jpg|IMG_002.jpg|IMG_003.jpg
"""

from pathlib import Path

from openpyxl import load_workbook

from app.imports.dto.product_import_dto import (
    ProductImportDTO,
)

from app.utils.price_parser import (
    parse_price_value,
)


class XlsxParser:

    def __init__(
        self,
        file_path: str | Path,
    ):
        self.file_path = Path(
            file_path
        )

    def parse(
        self,
    ) -> list[ProductImportDTO]:

        workbook = load_workbook(
            filename=self.file_path,
            data_only=True,
        )

        sheet = workbook["Товары"]

        products: list[
            ProductImportDTO
        ] = []

        headers = [
            cell.value
            for cell in sheet[1]
        ]

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        ):

            row_data = dict(
                zip(
                    headers,
                    row,
                )
            )

            # ==================================
            # ОБЯЗАТЕЛЬНЫЕ ПОЛЯ
            # ==================================

            sku = str(
                row_data.get(
                    "sku",
                    "",
                )
            ).strip()

            if not sku:
                continue

            title = str(
                row_data.get(
                    "title",
                    "",
                )
            ).strip()

            if not title:
                continue

            # ==================================
            # ЦЕНА
            # ==================================

            price_from_value = None
            price_to_value = None

            raw_price = row_data.get(
                "price"
            )

            if raw_price not in (
                None,
                "",
            ):

                (
                    price_from_value,
                    price_to_value,
                ) = parse_price_value(
                    str(raw_price)
                )

            # ==================================
            # DTO
            # ==================================

            product = (
                ProductImportDTO(
                    sku=sku,

                    title=title,

                    short_description=
                    self._clean_string(
                        row_data.get(
                            "short_description"
                        )
                    ),

                    full_description=
                    self._clean_string(
                        row_data.get(
                            "full_description"
                        )
                    ),

                    content_language=
                    self._clean_string(
                        row_data.get(
                            "content_language"
                        )
                    ),

                    category_slug=
                    self._clean_string(
                        row_data.get(
                            "category_slug"
                        )
                    ),

                    subcategory_slug=
                    self._clean_string(
                        row_data.get(
                            "subcategory_slug"
                        )
                    ),

                    brand_name=
                    self._clean_string(
                        row_data.get(
                            "brand_name"
                        )
                    ),

                    manufacturer_name=
                    self._clean_string(
                        row_data.get(
                            "manufacturer_name"
                        )
                    ),

                    manufacturer_sku=
                    self._clean_string(
                        row_data.get(
                            "manufacturer_sku"
                        )
                    ),

                    barcode=
                    self._clean_string(
                        row_data.get(
                            "barcode"
                        )
                    ),

                    serial_number=
                    self._clean_string(
                        row_data.get(
                            "serial_number"
                        )
                    ),

                    condition=
                    self._clean_string(
                        row_data.get(
                            "condition"
                        )
                    ),

                    price_type=
                    self._clean_string(
                        row_data.get(
                            "price_type"
                        )
                    ),

                    currency=
                    self._clean_string(
                        row_data.get(
                            "currency"
                        )
                    ),

                    price_from_value=
                    price_from_value,

                    price_to_value=
                    price_to_value,

                    quantity=
                    self._to_int(
                        row_data.get(
                            "quantity"
                        ),
                        default=1,
                    ),

                    weight_g=
                    self._to_int_nullable(
                        row_data.get(
                            "weight_g"
                        )
                    ),

                    dimensions=
                    self._clean_string(
                        row_data.get(
                            "dimensions"
                        )
                    ),

                    location_area=
                    self._clean_string(
                        row_data.get(
                            "location_area"
                        )
                    ),

                    video_url=
                    self._clean_string(
                        row_data.get(
                            "video_url"
                        )
                    ),

                    attributes_json=
                    self._clean_string(
                        row_data.get(
                            "attributes_json"
                        )
                    ),

                    photo_files=
                    self._clean_string(
                        row_data.get(
                            "photo_files"
                        )
                    ),
                )
            )

            products.append(
                product
            )

        return products

    @staticmethod
    def _clean_string(
        value,
    ) -> str | None:

        if value is None:
            return None

        value = str(value).strip()

        if not value:
            return None

        return value

    @staticmethod
    def _to_int(
        value,
        default: int,
    ) -> int:

        if value in (
            None,
            "",
        ):
            return default

        return int(value)

    @staticmethod
    def _to_int_nullable(
        value,
    ) -> int | None:

        if value in (
            None,
            "",
        ):
            return None

        return int(value)



# #
# # Парсер XLSX для TELESHOP.
# #
# # Поддерживаемый лист:
# #
# #     Товары
# #
# # Поддерживаемые колонки:
# #
# #     sku
# #     title
# #
# #     short_description
# #     full_description
# #
# #     category_slug
# #     brand_name
# #
# #     price_type
# #     currency
# #     price
# #
# #     quantity
# #     weight_g
# #     dimensions
# #
# # Важно:
# #
# # price:
# #
# #     500
# #     500-2000
# #     10.5
# #     10,5
# #     0.55-0.83
# #
# # разбирается через:
# #
# #     parse_price_value()
# #
#
# from pathlib import Path
#
# from openpyxl import load_workbook
#
# from app.imports.dto.product_import_dto import (
#     ProductImportDTO,
# )
#
# from app.utils.price_parser import (
#     parse_price_value,
# )
#
#
# class XlsxParser:
#
#     def __init__(
#         self,
#         file_path: str | Path,
#     ):
#         self.file_path = Path(
#             file_path
#         )
#
#     def parse(
#         self,
#     ) -> list[ProductImportDTO]:
#
#         workbook = load_workbook(
#             filename=self.file_path,
#             data_only=True,
#         )
#
#         sheet = workbook["Товары"]
#
#         products: list[
#             ProductImportDTO
#         ] = []
#
#         headers = [
#             cell.value
#             for cell in sheet[1]
#         ]
#
#         for row in sheet.iter_rows(
#             min_row=2,
#             values_only=True,
#         ):
#
#             row_data = dict(
#                 zip(
#                     headers,
#                     row,
#                 )
#             )
#
#             sku = str(
#                 row_data.get(
#                     "sku",
#                     "",
#                 )
#             ).strip()
#
#             if not sku:
#                 continue
#
#             title = str(
#                 row_data.get(
#                     "title",
#                     "",
#                 )
#             ).strip()
#
#             if not title:
#                 continue
#
#             price_from_value = None
#             price_to_value = None
#
#             raw_price = row_data.get(
#                 "price"
#             )
#
#             if raw_price not in (
#                 None,
#                 "",
#             ):
#
#                 (
#                     price_from_value,
#                     price_to_value,
#                 ) = parse_price_value(
#                     str(raw_price)
#                 )
#
#             product = (
#                 ProductImportDTO(
#                     sku=sku,
#
#                     title=title,
#
#                     short_description=
#                     row_data.get(
#                         "short_description"
#                     ),
#
#                     full_description=
#                     row_data.get(
#                         "full_description"
#                     ),
#
#                     category_slug=
#                     row_data.get(
#                         "category_slug"
#                     ),
#
#                     brand_name=
#                     row_data.get(
#                         "brand_name"
#                     ),
#
#                     price_type=
#                     row_data.get(
#                         "price_type"
#                     ),
#
#                     currency=
#                     row_data.get(
#                         "currency"
#                     ),
#
#                     price_from_value=
#                     price_from_value,
#
#                     price_to_value=
#                     price_to_value,
#
#                     quantity=
#                     row_data.get(
#                         "quantity"
#                     )
#                     or 1,
#
#                     weight_g=
#                     row_data.get(
#                         "weight_g"
#                     ),
#
#                     dimensions=
#                     row_data.get(
#                         "dimensions"
#                     ),
#                 )
#             )
#
#             products.append(
#                 product
#             )
#
#         return products
#
#
#
#
#
# # # app/imports/parsers/xlsx_parser.py
# #
# # from pathlib import Path
# #
# # from openpyxl import load_workbook
# #
# # from app.imports.dto.product_import_dto import (
# #     ProductImportDTO,
# # )
# #
# #
# # class XlsxParser:
# #     """
# #     Парсер XLSX файла импорта товаров.
# #
# #     Поддерживается лист:
# #
# #         Товары
# #
# #     Каждая строка листа соответствует
# #     одному товару.
# #     """
# #
# #     def __init__(
# #         self,
# #         file_path: str | Path,
# #     ):
# #         self.file_path = Path(file_path)
# #
# #     def parse(
# #         self,
# #     ) -> list[ProductImportDTO]:
# #         """
# #         Читает XLSX файл и возвращает
# #         список DTO товаров.
# #         """
# #
# #         workbook = load_workbook(
# #             filename=self.file_path,
# #             data_only=True,
# #         )
# #
# #         sheet = workbook["Товары"]
# #
# #         products: list[ProductImportDTO] = []
# #
# #         headers = [
# #             cell.value
# #             for cell in sheet[1]
# #         ]
# #
# #         for row in sheet.iter_rows(
# #             min_row=2,
# #             values_only=True,
# #         ):
# #
# #             row_data = dict(
# #                 zip(headers, row)
# #             )
# #
# #             photos = self._parse_list(
# #                 row_data.get("photos")
# #             )
# #
# #             tags = self._parse_list(
# #                 row_data.get("tags")
# #             )
# #
# #             product = ProductImportDTO(
# #                 sku=str(
# #                     row_data.get(
# #                         "sku",
# #                         "",
# #                     )
# #                 ).strip(),
# #
# #                 manufacturer_name=row_data.get(
# #                     "manufacturer_name"
# #                 ),
# #
# #                 manufacturer_sku=row_data.get(
# #                     "manufacturer_sku"
# #                 ),
# #
# #                 title=str(
# #                     row_data.get(
# #                         "title",
# #                         "",
# #                     )
# #                 ).strip(),
# #
# #                 category_path=row_data.get(
# #                     "category_path"
# #                 ),
# #
# #                 brand=row_data.get(
# #                     "brand"
# #                 ),
# #
# #                 condition=row_data.get(
# #                     "condition"
# #                 ) or "USED",
# #
# #                 status=row_data.get(
# #                     "status"
# #                 ),
# #
# #                 price_value=row_data.get(
# #                     "price_value"
# #                 ),
# #
# #                 currency=row_data.get(
# #                     "currency"
# #                 ) or "UAH",
# #
# #                 quantity=row_data.get(
# #                     "quantity"
# #                 ) or 1,
# #
# #                 weight_g=row_data.get(
# #                     "weight_g"
# #                 ),
# #
# #                 dimensions=row_data.get(
# #                     "dimensions"
# #                 ),
# #
# #                 short_description=row_data.get(
# #                     "short_description"
# #                 ),
# #
# #                 full_description=row_data.get(
# #                     "full_description"
# #                 ),
# #
# #                 video_url=row_data.get(
# #                     "video_url"
# #                 ),
# #
# #                 photos=photos,
# #
# #                 sort_priority=row_data.get(
# #                     "sort_priority"
# #                 ) or 0,
# #
# #                 warranty_days=row_data.get(
# #                     "warranty_days"
# #                 ),
# #
# #                 tags=tags,
# #
# #                 is_published=bool(
# #                     row_data.get(
# #                         "is_published",
# #                         True,
# #                     )
# #                 ),
# #             )
# #
# #             products.append(
# #                 product
# #             )
# #
# #         return products
# #
# #     @staticmethod
# #     def _parse_list(
# #         value: str | None,
# #     ) -> list[str]:
# #         """
# #         Преобразует строку:
# #
# #             photo1.jpg|photo2.jpg
# #
# #         в список:
# #
# #             [
# #                 "photo1.jpg",
# #                 "photo2.jpg",
# #             ]
# #         """
# #
# #         if not value:
# #             return []
# #
# #         return [
# #             item.strip()
# #             for item in str(value).split("|")
# #             if item.strip()
# #         ]